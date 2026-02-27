"""
update_excel.py
Manages the jobs_tracker.xlsx Excel file using openpyxl.

Columns (fixed order):
  job_id | title | company | location | apply_url | match_score |
  strengths | gaps | keywords_missing | tailored_resume_path |
  status | scraped_at | applied_at | notes
"""

import logging
import shutil
import tempfile
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

log = logging.getLogger(__name__)

COLUMNS = [
    "job_id", "title", "company", "location", "apply_url",
    "match_score", "strengths", "gaps", "keywords_missing",
    "tailored_resume_path", "status", "scraped_at", "applied_at", "notes",
]

# Valid statuses
STATUS_PENDING  = "pending"
STATUS_APPLIED  = "applied"
STATUS_FAILED   = "failed"
STATUS_SKIPPED  = "skipped"
STATUS_MANUAL   = "manual_required"   # CAPTCHA or other blocker — apply manually

STATUS_COLORS = {
    STATUS_PENDING:  "FFF9C4",   # yellow
    STATUS_APPLIED:  "C8E6C9",   # green
    STATUS_FAILED:   "FFCDD2",   # red
    STATUS_SKIPPED:  "E0E0E0",   # grey
    STATUS_MANUAL:   "FFE0B2",   # orange
}


def _safe_save(wb: openpyxl.Workbook, path: str) -> str:
    """
    Save workbook to path. If the file is locked (e.g. open in Excel),
    save to a timestamped fallback alongside it and warn the user.
    Returns the actual path saved to.
    """
    # Write to a temp file first, then atomically replace the target
    p = Path(path)
    try:
        with tempfile.NamedTemporaryFile(
            dir=p.parent, suffix=".xlsx", delete=False
        ) as tmp:
            tmp_path = tmp.name
        wb.save(tmp_path)
        # Atomic replace — works even if target exists (not if it's locked)
        try:
            shutil.move(tmp_path, path)
            return path
        except PermissionError:
            # Target is locked (open in Excel). Keep the temp file as fallback.
            fallback = str(p.parent / f"{p.stem}_{datetime.now().strftime('%H%M%S')}{p.suffix}")
            shutil.move(tmp_path, fallback)
            log.warning(
                f"jobs_tracker.xlsx is open in Excel — saved to {fallback} instead. "
                "Close Excel and merge manually, or close Excel before running the pipeline."
            )
            return fallback
    except Exception as exc:
        Path(tmp_path).unlink(missing_ok=True)
        raise exc


def _col_index(name: str) -> int:
    return COLUMNS.index(name) + 1  # 1-based


def init_tracker(path: str) -> None:
    """Create the Excel tracker with a styled header row if it doesn't exist."""
    p = Path(path)
    if p.exists():
        return
    p.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jobs"

    header_fill = PatternFill("solid", fgColor="1565C0")
    header_font = Font(bold=True, color="FFFFFF")

    for i, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i, value=col.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Set column widths
    widths = {
        "job_id": 14, "title": 35, "company": 25, "location": 20,
        "apply_url": 45, "match_score": 12, "strengths": 40,
        "gaps": 40, "keywords_missing": 40, "tailored_resume_path": 45,
        "status": 16, "scraped_at": 22, "applied_at": 22, "notes": 40,
    }
    for col_name, width in widths.items():
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(_col_index(col_name))
        ].width = width

    ws.freeze_panes = "A2"
    _safe_save(wb, str(p))
    log.info(f"Created tracker: {p}")


def job_exists(path: str, job_id: str) -> bool:
    """Return True if job_id is already in column A."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] == job_id:
                return True
    except Exception:
        pass
    return False


def _list_to_str(value) -> str:
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    return str(value) if value is not None else ""


def add_job(path: str, job_dict: dict) -> None:
    """Append a new job row. Lists are stored as semicolon-separated strings."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    status = job_dict.get("status", STATUS_PENDING)
    fill_color = STATUS_COLORS.get(status, "FFFFFF")
    fill = PatternFill("solid", fgColor=fill_color)

    row_data = {
        "job_id":              job_dict.get("job_id", ""),
        "title":               job_dict.get("title", ""),
        "company":             job_dict.get("company", ""),
        "location":            job_dict.get("location", ""),
        "apply_url":           job_dict.get("apply_url", ""),
        "match_score":         job_dict.get("match_score", ""),
        "strengths":           _list_to_str(job_dict.get("strengths", [])),
        "gaps":                _list_to_str(job_dict.get("gaps", [])),
        "keywords_missing":    _list_to_str(job_dict.get("keywords_missing", [])),
        "tailored_resume_path": job_dict.get("tailored_resume_path", ""),
        "status":              status,
        "scraped_at":          job_dict.get("scraped_at", ""),
        "applied_at":          "",
        "notes":               job_dict.get("notes", ""),
    }

    new_row = ws.max_row + 1
    for col_name, value in row_data.items():
        cell = ws.cell(row=new_row, column=_col_index(col_name), value=value)
        cell.fill = fill

    _safe_save(wb, path)
    log.debug(f"Added job {row_data['job_id']} to tracker")


def _find_workbook_with_job(path: str, job_id: str) -> str:
    """
    Return the path of the xlsx file that contains job_id.
    Checks the primary path first, then any timestamped fallback files in the same dir.
    """
    candidates = [path]
    p = Path(path)
    # Fallbacks look like jobs_tracker_HHMMSS.xlsx
    candidates += sorted(p.parent.glob(f"{p.stem}_*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    for candidate in candidates:
        try:
            if job_exists(str(candidate), job_id):
                return str(candidate)
        except Exception:
            continue
    return path  # default to primary even if not found


def update_status(
    path: str,
    job_id: str,
    status: str,
    notes: str = "",
) -> bool:
    """
    Find row by job_id (searching primary + any fallback files), update status/applied_at/notes.
    Returns True if found and updated.
    """
    actual_path = _find_workbook_with_job(path, job_id)
    if actual_path != path:
        log.info(f"update_status: found job {job_id} in fallback file {actual_path}")
    wb = openpyxl.load_workbook(actual_path)
    ws = wb.active

    id_col  = _col_index("job_id")
    st_col  = _col_index("status")
    at_col  = _col_index("applied_at")
    nt_col  = _col_index("notes")

    fill_color = STATUS_COLORS.get(status, "FFFFFF")
    fill = PatternFill("solid", fgColor=fill_color)

    for row in ws.iter_rows(min_row=2):
        if row[id_col - 1].value == job_id:
            row[st_col - 1].value = status
            row[at_col - 1].value = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            row[nt_col - 1].value = notes
            # Apply color to entire row
            for cell in row:
                cell.fill = fill
            _safe_save(wb, actual_path)
            log.info(f"Updated job {job_id} → {status}")
            return True

    log.warning(f"job_id {job_id} not found in tracker")
    return False


def get_pending_jobs(path: str) -> list[dict]:
    """Return all rows with status='pending' as list of dicts."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
    except FileNotFoundError:
        return []

    pending = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(COLUMNS, row))
        if row_dict.get("status") == STATUS_PENDING:
            pending.append(row_dict)
    return pending
